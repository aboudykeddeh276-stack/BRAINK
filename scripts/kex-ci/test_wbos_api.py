#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import subprocess
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[2]
SERVER = ROOT / "modules" / "kex_wbos" / "server.py"
OPENAPI = ROOT / "openapi" / "kex-unified-api.yaml"
LEDGER = ROOT / "reports" / "kex-wbos" / "proof-ledger.jsonl"
RUNTIME_BOOKS = ROOT / "runtime" / "workbooks"
BASE = "http://127.0.0.1:8765"


def get(path: str):
    with urllib.request.urlopen(BASE + path, timeout=5) as response:
        return response.status, response.headers.get("Content-Type", ""), response.read()


def get_allow_404(path: str):
    try:
        return get(path)
    except urllib.error.HTTPError as exc:
        return exc.code, exc.headers.get("Content-Type", ""), exc.read()


def post_multipart(path: str, fields: dict[str, tuple[str, bytes]]):
    boundary = "----KEXBoundary7MA4YWxkTrZu0gW"
    chunks: list[bytes] = []
    for name, (filename, payload) in fields.items():
        chunks.extend([
            f"--{boundary}\r\n".encode(),
            f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'.encode(),
            b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n",
            payload,
            b"\r\n",
        ])
    chunks.append(f"--{boundary}--\r\n".encode())
    body = b"".join(chunks)
    req = urllib.request.Request(BASE + path, data=body, method="POST", headers={"Content-Type": f"multipart/form-data; boundary={boundary}"})
    with urllib.request.urlopen(req, timeout=10) as response:
        return response.status, response.headers.get("Content-Type", ""), response.read()


def source_workbook_bytes() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    def add(name: str, headers: list[str], rows: list[list[object]]):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    add("CORE_LATTICE", ["address", "value", "type", "state"], [["K.0", 1, "ROOT", "ACTIVE"], ["K.1", 0.297, "RESONANCE", "ACTIVE"]])
    add("ROOT_MATRIX", ["id", "v1", "v2", "v3"], [[1, -3, 1, 3], [2, -2, 1, 2]])
    add("GLOBAL_INDEX", ["timestamp", "term", "value"], [["2026-09-01T00:00:00+00:00", "alpha", 1], ["2026-09-01T01:00:00+00:00", "beta", 2]])
    add("BENCHMARKS", ["operation", "timeSec"], [["hydrate", 0.12], ["resolve", 0.03]])
    add("HYPER_CORES", ["id", "model", "status", "capacity"], [["HC-1", "logical", "ACTIVE", "1 lane"]])
    add("SERVERS", ["id", "hostname", "status", "cpuCores", "memory"], [["SRV-1", "wbos-local", "ACTIVE", 4, "8 GB"]])
    add("STORAGE_DEVICES", ["id", "type", "capacity", "status"], [["STO-1", "virtual", "100 TB", "CONTRACT"]])
    add("LOGS", ["timestamp", "operation", "details"], [["2026-09-01T00:30:00+00:00", "boot", "test"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def second_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "KEX_DNA"
    ws.append(["index", "identity", "symbol", "role"])
    ws.append([1, "Hydrogen", "H", "PRIMARY_IGNITION"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> int:
    if LEDGER.exists():
        LEDGER.unlink()
    if RUNTIME_BOOKS.exists():
        for path in RUNTIME_BOOKS.glob("*.xls*"):
            path.unlink()

    spec = OPENAPI.read_text(encoding="utf-8")
    for required in ["openapi: 3.1.0", "/core-lattice:", "/global-index:", "/root-matrix/statistics:", "/activate-workbook:", "/workbooks/apply:", "/api/health:", "/api/proof-ledger:"]:
        require(required in spec, f"KEX Systems OpenAPI surface missing: {required}")

    proc = subprocess.Popen([sys.executable, str(SERVER)], cwd=ROOT)
    try:
        deadline = time.time() + 10
        while True:
            try:
                status, _, _ = get("/api/health")
                if status == 200:
                    break
            except Exception:
                pass
            if time.time() >= deadline:
                raise AssertionError("KEX Systems API did not become reachable")
            time.sleep(0.1)

        _, _, raw = get("/core-lattice")
        before = json.loads(raw)
        require(before["state"] == "SOURCE_NOT_RESIDENT", "pre-activation source boundary failed")

        first = source_workbook_bytes()
        second = second_workbook_bytes()
        status, _, raw = post_multipart("/activate-workbook", {"workbook": ("kex-source-test.xlsx", first)})
        activation = json.loads(raw)
        require(status == 200, "workbook activation failed")
        require(activation["parse_state"] == "STORED_AND_INDEXED", "activated workbook was not indexed")
        require("CORE_LATTICE" in activation["sheets"], "activation did not expose workbook sheet registry")

        _, _, raw = get("/core-lattice")
        lattice = json.loads(raw)
        require(lattice["state"] == "RESIDENT", "core lattice did not promote after source activation")
        require(lattice["row_count"] == 2, "core lattice row readback wrong")
        require(lattice["rows"][1]["value"] == 0.297, "resonance value was not preserved")

        _, _, raw = get("/root-matrix/statistics")
        stats = json.loads(raw)
        require(stats["state"] == "COMPUTED_FROM_RESIDENT_ROWS", "root matrix statistics did not compute from resident source")
        require(stats["count"] == 6 and stats["min"] == -3 and stats["max"] == 3, "root matrix statistics incorrect")

        _, _, raw = get("/global-index?timestampFrom=2026-09-01T00:30:00%2B00:00")
        global_index = json.loads(raw)
        require(global_index["row_count"] == 1 and global_index["rows"][0]["term"] == "beta", "global index timestamp filtering failed")

        _, _, raw = get("/benchmarks?operation=resolve")
        benchmarks = json.loads(raw)
        require(benchmarks["row_count"] == 1 and benchmarks["rows"][0]["operation"] == "resolve", "benchmark operation filtering failed")

        status, _, raw = get_allow_404("/hyper-cores/HC-1")
        require(status == 200 and json.loads(raw)["id"] == "HC-1", "hyper-core object lookup failed")
        status, _, raw = get_allow_404("/servers/SRV-1")
        require(status == 200 and json.loads(raw)["hostname"] == "wbos-local", "server object lookup failed")
        status, _, raw = get_allow_404("/storage-devices/STO-1")
        require(status == 200 and json.loads(raw)["capacity"] == "100 TB", "storage object lookup failed")

        _, _, raw = get("/storage-summary")
        storage = json.loads(raw)
        require(storage["state"] == "COMPUTED_FROM_RESIDENT_ROWS", "storage summary did not derive from source")
        require(storage["categories"][0]["totalCapacityTb"] == 100.0, "storage capacity summary incorrect")

        _, _, raw = get("/system-summary")
        system = json.loads(raw)
        require(system["state"] == "COMPUTED_FROM_RESIDENT_ROWS", "system summary did not derive from source")

        status, ctype, combined = post_multipart("/workbooks/apply", {"workbook1": ("a.xlsx", first), "workbook2": ("b.xlsx", second)})
        require(status == 200 and "spreadsheetml.sheet" in ctype, "workbook apply failed")
        wb = load_workbook(io.BytesIO(combined), data_only=False)
        require("A_CORE_LATTICE" in wb.sheetnames and "B_KEX_DNA" in wb.sheetnames, "combined workbook sheet materialisation failed")
        require(wb["B_KEX_DNA"]["D2"].value == "PRIMARY_IGNITION", "combined workbook value preservation failed")

        _, _, raw = get("/virtualization-metrics")
        metrics = json.loads(raw)
        require(metrics["state"] == "UNMEASURED", "virtualization metrics were fabricated")

        _, _, raw = get("/mesh")
        mesh = json.loads(raw)
        require(mesh["state"] == "LOCAL_REGISTRY_ONLY", "mesh boundary was over-promoted")

        _, _, raw = get("/api/proof-ledger")
        entries = json.loads(raw)["entries"]
        events = {entry["event"] for entry in entries}
        for event in ["WBOS_BOOT", "HEALTH_READ", "WORKBOOK_DATA_READ", "ROOT_MATRIX_STATISTICS", "ACTIVATE_WORKBOOK", "APPLY_WORKBOOKS"]:
            require(event in events, f"proof event missing: {event}")

        receipt = {
            "schema": "kex.systems.integration-test.v3",
            "state": "PASS",
            "source_transition": "SOURCE_NOT_RESIDENT -> RESIDENT",
            "activated_workbook": activation["path"],
            "activated_sheet_count": len(activation["sheets"]),
            "core_lattice_rows": lattice["row_count"],
            "root_matrix_statistics": stats,
            "global_index_filter": "PASS",
            "benchmark_filter": "PASS",
            "object_id_lookups": "PASS",
            "storage_summary": "PASS",
            "system_summary": "PASS",
            "workbook_apply": "executed_and_read_back",
            "proof_entries": len(entries),
            "claim_boundary": "This proves local workbook activation, resident sheet discovery/parsing, filtered data retrieval, computed summaries/statistics, object lookup, workbook merge/readback and proof mutation in this execution context. It does not prove unavailable historical workbooks are resident, public deployment, external mesh connectivity, browser IndexedDB execution, or physical storage provisioning."
        }
        out = ROOT / "reports" / "kex-wbos" / "integration.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(receipt, indent=2) + "\n", encoding="utf-8")
        print(json.dumps(receipt, indent=2))
        return 0
    finally:
        proc.terminate()
        try:
            proc.wait(timeout=3)
        except subprocess.TimeoutExpired:
            proc.kill()
            proc.wait(timeout=3)


if __name__ == "__main__":
    raise SystemExit(main())
