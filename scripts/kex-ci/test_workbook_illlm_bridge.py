#!/usr/bin/env python3
from __future__ import annotations

import tempfile
from pathlib import Path
import sys

from openpyxl import Workbook

BASE = Path(__file__).resolve().parents[2]
MODULES = BASE / "modules" / "kex_wbos"
sys.path.insert(0, str(MODULES))

from illlm_recursive_runtime import RecursiveILLLMRuntime  # noqa: E402
from workbook_illlm_bridge import hydrate_workbook_into_illlm  # noqa: E402


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        path = root / "demo.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "MATH"
        ws["A1"] = 2
        ws["A2"] = 3
        ws["A3"] = "=A1+A2"
        ws["B1"] = "=SUM(A1:A3)"
        wb.save(path)
        wb.close()

        runtime = RecursiveILLLMRuntime()
        receipt = hydrate_workbook_into_illlm(runtime, path)
        assert receipt["status"] == "HYDRATED"
        assert receipt["sheetCount"] == 1
        assert receipt["semanticObjectCount"] >= 4
        assert receipt["dependencyEdgesAdded"] >= 3
        assert receipt["runtimeGraphHash"] == runtime.graph_hash()
        assert any(node.role == "WORKBOOK_FORMULA" for node in runtime.nodes.values())
        assert any(edge.relation == "WORKBOOK_FEEDS" for edges in runtime.edges_out.values() for edge in edges)
        assert any(node.carrier == path.as_posix() for node in runtime.nodes.values())

    print("PASS workbook→IL-LLM bridge")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
