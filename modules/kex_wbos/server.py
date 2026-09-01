#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import time
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from workbook_api import (
    DATASETS,
    activation_receipt,
    dataset_response,
    discover_workbooks,
    root_matrix_statistics,
    storage_summary,
    system_summary,
    virtualization_metrics,
)

OS_NAME = "KEX-WBOS Cascade OS"
PORT = 8765
ROOT_URI = "KEX://ROOT/OS"
CASCADE_ORDER = [
    "MOUNT", "VERIFY_ALL", "HYDRATE_INDEXEDDB", "RESOLVE_KEX_URI",
    "PROJECT_UI", "DISPATCH_TRIGGER", "MUTATE_STATE", "WRITEBACK",
    "PROOF_COMMIT", "REHYDRATE_ON_FAILURE",
]

BASE = Path(__file__).resolve().parents[2]
LEDGER = BASE / "reports" / "kex-wbos" / "proof-ledger.jsonl"

SERVICES = [
    {"service": "apex", "route": "/", "port": PORT, "role": "K.0 apex projection"},
    {"service": "hybrid-os", "route": "/hybrid-os", "port": PORT, "role": "hybrid OS projection"},
    {"service": "health", "route": "/api/health", "port": PORT, "role": "runtime health"},
    {"service": "resolver", "route": "/api/resolve", "port": PORT, "role": "KEX URI resolution"},
    {"service": "proof-ledger", "route": "/api/proof-ledger", "port": PORT, "role": "proof readback"},
    {"service": "workbook-api", "route": "/core-lattice", "port": PORT, "role": "workbook source adapter"},
]

ROUTES = [
    {"host": "127.0.0.1", "path": "/", "target": "wbos.apex", "kex": ROOT_URI},
    {"host": "127.0.0.1", "path": "/hybrid-os", "target": "wbos.hybrid", "kex": "KEX://ROOT/HYBRID_OS"},
    {"host": "127.0.0.1", "path": "/api/health", "target": "wbos.health", "kex": "KEX://ROOT/OS/HEALTH"},
    {"host": "127.0.0.1", "path": "/api/services", "target": "wbos.services", "kex": "KEX://ROOT/OS/SERVICES"},
    {"host": "127.0.0.1", "path": "/api/routes", "target": "wbos.routes", "kex": "KEX://ROOT/OS/ROUTES"},
    {"host": "127.0.0.1", "path": "/mesh", "target": "braink.mesh", "kex": "KEX://BRAINK/MESH"},
    {"host": "127.0.0.1", "path": "/cascade", "target": "wbos.cascade", "kex": "KEX://ROOT/OS/CASCADE"},
    {"host": "127.0.0.1", "path": "/core-lattice", "target": "wbos.workbook.core-lattice", "kex": "KEX://ROOT/WORKBOOK/CORE_LATTICE"},
    {"host": "127.0.0.1", "path": "/genome-store", "target": "wbos.workbook.genome-store", "kex": "KEX://ROOT/WORKBOOK/GENOME_STORE"},
]
ROUTE_BY_KEX = {row["kex"].upper(): row for row in ROUTES}


def append_proof(event: str, target: str, value: str, actor: str = "kex-wbos") -> dict:
    entry = {"ts": time.time(), "event": event, "target": target, "value": value, "actor": actor}
    LEDGER.parent.mkdir(parents=True, exist_ok=True)
    with LEDGER.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, sort_keys=True) + "\n")
    return entry


def read_proofs(limit: int = 100) -> list[dict]:
    if not LEDGER.exists():
        return []
    rows = []
    for line in LEDGER.read_text(encoding="utf-8").splitlines():
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return rows[-limit:]


def resolve_kex(uri: str) -> dict:
    normalized = (uri or ROOT_URI).strip().upper()
    route = ROUTE_BY_KEX.get(normalized)
    if route:
        return {"found": True, "type": "route", **route}
    return {"found": False, "type": "unresolved", "target": None, "kex": normalized}


def mesh_status() -> dict:
    return {"braink_mesh": "configured", "nodes": ["wbos-local"], "state": "LOCAL_REGISTRY_ONLY"}


def cascade_info() -> dict:
    return {"root": ROOT_URI, "order": CASCADE_ORDER, "children": [s["service"] for s in SERVICES]}


def parse_multipart(content_type: str, body: bytes) -> dict[str, tuple[str, bytes]]:
    marker = "boundary="
    if marker not in content_type:
        return {}
    boundary = content_type.split(marker, 1)[1].strip().strip('"').encode()
    result: dict[str, tuple[str, bytes]] = {}
    for part in body.split(b"--" + boundary):
        if b"\r\n\r\n" not in part:
            continue
        headers, payload = part.split(b"\r\n\r\n", 1)
        payload = payload.rstrip(b"\r\n-")
        header_text = headers.decode("utf-8", errors="replace")
        if "Content-Disposition:" not in header_text:
            continue
        name = None
        filename = "upload.bin"
        for token in header_text.replace(";", "\n").splitlines():
            token = token.strip()
            if token.startswith("name="):
                name = token.split("=", 1)[1].strip('"')
            elif token.startswith("filename="):
                filename = token.split("=", 1)[1].strip('"')
        if name:
            result[name] = (filename, payload)
    return result


def combine_xlsx(first: tuple[str, bytes], second: tuple[str, bytes]) -> bytes:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for workbook combination") from exc
    output = Workbook()
    output.remove(output.active)
    for prefix, pair in (("A", first), ("B", second)):
        _, raw = pair
        source = load_workbook(io.BytesIO(raw), data_only=False)
        for sheet in source.worksheets:
            title = f"{prefix}_{sheet.title}"[:31]
            ws = output.create_sheet(title)
            for row in sheet.iter_rows():
                for cell in row:
                    ws[cell.coordinate] = cell.value
    buf = io.BytesIO()
    output.save(buf)
    return buf.getvalue()


class Handler(BaseHTTPRequestHandler):
    server_version = "KEX-Unified/2.0"

    def _json(self, payload: object, status: int = 200) -> None:
        body = json.dumps(payload, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _binary(self, payload: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def _html(self, title: str, body_text: str) -> None:
        body = f"<!doctype html><html><head><title>{title}</title></head><body><main><h1>{title}</h1><p>{body_text}</p></main></body></html>".encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt: str, *args) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        dataset_key = path.lstrip("/")

        if path == "/":
            append_proof("PROJECT_UI", "wbos.apex", "rendered")
            return self._html("KEX K.0 Apex", "KEX Unified API and WBOS local apex projection.")
        if path == "/hybrid-os":
            append_proof("PROJECT_UI", "wbos.hybrid", "rendered")
            return self._html("KEX Hybrid OS", "Hybrid OS local projection.")
        if path == "/api/health":
            proof = append_proof("HEALTH_READ", ROOT_URI, "ok")
            return self._json({"status": "ok", "os": OS_NAME, "cascade_order": CASCADE_ORDER, "proof": proof, "workbooks": discover_workbooks()})
        if path == "/api/services":
            append_proof("SERVICE_LIST_READ", ROOT_URI, str(len(SERVICES)))
            return self._json({"services": SERVICES})
        if path == "/api/routes":
            append_proof("ROUTE_LIST_READ", ROOT_URI, str(len(ROUTES)))
            return self._json({"routes": ROUTES})
        if path == "/api/resolve":
            uri = parse_qs(parsed.query).get("uri", [ROOT_URI])[0]
            result = resolve_kex(uri)
            append_proof("RESOLVE_KEX_URI", uri, "found" if result["found"] else "unresolved")
            return self._json({"uri": uri, "result": result})
        if path == "/mesh":
            append_proof("MESH_STATUS_READ", "braink.mesh", "local-registry")
            return self._json(mesh_status())
        if path == "/cascade":
            append_proof("CASCADE_READ", ROOT_URI, str(len(CASCADE_ORDER)))
            return self._json(cascade_info())
        if path == "/api/proof-ledger":
            return self._json({"entries": read_proofs(100)})
        if dataset_key in DATASETS:
            result = dataset_response(dataset_key)
            append_proof("WORKBOOK_DATA_READ", dataset_key, result["state"])
            return self._json(result)
        if path == "/root-matrix/statistics":
            rows = dataset_response("root-matrix")["rows"]
            result = root_matrix_statistics(rows)
            append_proof("ROOT_MATRIX_STATISTICS", "root-matrix", result["state"])
            return self._json(result)
        if path == "/storage-summary":
            rows = dataset_response("storage-devices")["rows"]
            return self._json(storage_summary(rows))
        if path == "/system-summary":
            return self._json(system_summary(
                dataset_response("hyper-cores")["rows"],
                dataset_response("servers")["rows"],
                dataset_response("storage-devices")["rows"],
            ))
        if path == "/virtualization-metrics":
            return self._json(virtualization_metrics())
        if path.startswith("/hyper-cores/"):
            return self._json({"state": "SOURCE_NOT_RESIDENT", "id": path.rsplit("/", 1)[-1]}, 404)
        if path.startswith("/servers/"):
            return self._json({"state": "SOURCE_NOT_RESIDENT", "id": path.rsplit("/", 1)[-1]}, 404)
        if path.startswith("/storage-devices/"):
            return self._json({"state": "SOURCE_NOT_RESIDENT", "id": path.rsplit("/", 1)[-1]}, 404)
        return self._json({"error": "not_found", "path": path}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length)
        parts = parse_multipart(self.headers.get("Content-Type", ""), body)

        if parsed.path == "/activate-workbook":
            if "workbook" not in parts:
                return self._json({"error": "workbook_required"}, 400)
            filename, payload = parts["workbook"]
            receipt = activation_receipt(filename, payload)
            append_proof("ACTIVATE_WORKBOOK", receipt["path"], receipt["sha256"])
            return self._json(receipt)

        if parsed.path == "/workbooks/apply":
            if "workbook1" not in parts or "workbook2" not in parts:
                return self._json({"error": "workbook1_and_workbook2_required"}, 400)
            try:
                combined = combine_xlsx(parts["workbook1"], parts["workbook2"])
            except Exception as exc:
                return self._json({"error": "workbook_combine_failed", "detail": str(exc)}, 500)
            append_proof("APPLY_WORKBOOKS", "combined.xlsx", str(len(combined)))
            return self._binary(combined, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "kex-combined.xlsx")

        return self._json({"error": "not_found", "path": parsed.path}, 404)


def serve(host: str = "127.0.0.1", port: int = PORT) -> None:
    append_proof("WBOS_BOOT", ROOT_URI, f"{host}:{port}")
    ThreadingHTTPServer((host, port), Handler).serve_forever()


if __name__ == "__main__":
    serve()
