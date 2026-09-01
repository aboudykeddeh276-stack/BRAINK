#!/usr/bin/env python3
from __future__ import annotations

import json
import os
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

from hardening import atomic_write_text, canonical_json_bytes, sha256_bytes

MAX_POPULATED_CELLS = int(os.getenv("KEX_WORKBOOK_MAX_SEMANTIC_CELLS", "250000"))
MAX_DEPENDENCY_EDGES = int(os.getenv("KEX_WORKBOOK_MAX_DEPENDENCY_EDGES", "1000000"))
MAX_RANGE_EXPANSION = int(os.getenv("KEX_WORKBOOK_MAX_RANGE_EXPANSION", "4096"))


def _split_sheet_ref(value: str, current_sheet: str) -> tuple[str, str]:
    value = value.replace("$", "")
    if "!" in value:
        sheet, ref = value.rsplit("!", 1)
        sheet = sheet.strip("'")
        return sheet, ref
    return current_sheet, value


def _expand_dependency(value: str, current_sheet: str) -> tuple[list[str], dict[str, Any] | None]:
    sheet, ref = _split_sheet_ref(value, current_sheet)
    if ":" not in ref:
        return [f"{sheet}!{ref}"], None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except ValueError:
        return [], {"objectId": f"range://{sheet}!{ref}", "sheet": sheet, "range": ref, "kind": "RANGE_UNPARSED"}
    count = (max_col - min_col + 1) * (max_row - min_row + 1)
    if count > MAX_RANGE_EXPANSION:
        return [], {
            "objectId": f"range://{sheet}!{ref}",
            "sheet": sheet,
            "range": ref,
            "kind": "RANGE",
            "cellCount": count,
            "expanded": False,
        }
    cells = [
        f"{sheet}!{get_column_letter(col)}{row}"
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    ]
    return cells, {
        "objectId": f"range://{sheet}!{ref}",
        "sheet": sheet,
        "range": ref,
        "kind": "RANGE",
        "cellCount": count,
        "expanded": True,
    }


def _formula_dependencies(formula: str, current_sheet: str) -> tuple[list[str], list[dict[str, Any]]]:
    deps: list[str] = []
    ranges: list[dict[str, Any]] = []
    try:
        tokens = Tokenizer(formula).items
    except Exception:
        return deps, ranges
    for token in tokens:
        if token.type != "OPERAND" or token.subtype != "RANGE":
            continue
        value = token.value
        if not value or value.startswith("["):
            continue
        expanded, range_obj = _expand_dependency(value, current_sheet)
        if range_obj and range_obj not in ranges:
            ranges.append(range_obj)
        for item in expanded:
            if item not in deps:
                deps.append(item)
    return deps, ranges


def _strongly_connected_components(edges: list[dict[str, str]]) -> list[list[str]]:
    graph: dict[str, set[str]] = defaultdict(set)
    reverse: dict[str, set[str]] = defaultdict(set)
    for edge in edges:
        source, target = edge["from"], edge["to"]
        graph[source].add(target)
        graph.setdefault(target, set())
        reverse[target].add(source)
        reverse.setdefault(source, set())

    visited: set[str] = set()
    order: list[str] = []
    for start in sorted(graph):
        if start in visited:
            continue
        visited.add(start)
        stack: list[tuple[str, bool]] = [(start, False)]
        while stack:
            node, exiting = stack.pop()
            if exiting:
                order.append(node)
                continue
            stack.append((node, True))
            for nxt in sorted(graph[node], reverse=True):
                if nxt not in visited:
                    visited.add(nxt)
                    stack.append((nxt, False))

    assigned: set[str] = set()
    components: list[list[str]] = []
    for start in reversed(order):
        if start in assigned:
            continue
        component: list[str] = []
        stack = [start]
        assigned.add(start)
        while stack:
            node = stack.pop()
            component.append(node)
            for nxt in reverse[node]:
                if nxt not in assigned:
                    assigned.add(nxt)
                    stack.append(nxt)
        components.append(sorted(component))
    return components


def _stored_cells(ws) -> list[Any]:
    # openpyxl materializes resident cells in _cells; iter_rows() would traverse
    # the rectangular dimension and is unsafe for sparse sheets with distant cells.
    cells = list(getattr(ws, "_cells", {}).values())
    cells.sort(key=lambda cell: (cell.row, cell.column))
    return cells


def analyze_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=False, data_only=False)
    objects: list[dict[str, Any]] = []
    range_objects: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, str]] = []
    formulas = 0
    populated = 0
    sheet_count = len(wb.sheetnames)

    try:
        for ws in wb.worksheets:
            for cell in _stored_cells(ws):
                value = cell.value
                if value is None:
                    continue
                populated += 1
                if populated > MAX_POPULATED_CELLS:
                    raise ValueError("workbook_semantic_cell_limit_exceeded")
                address = f"{ws.title}!{cell.coordinate}"
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    deps, ranges = _formula_dependencies(value, ws.title)
                    for range_obj in ranges:
                        range_objects[range_obj["objectId"]] = range_obj
                    objects.append({
                        "objectId": f"cell://{address}",
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "kind": "FORMULA",
                        "formula": value,
                        "dependencies": deps,
                        "rangeDependencies": [item["objectId"] for item in ranges],
                    })
                    for dep in deps:
                        edges.append({"from": f"cell://{dep}", "to": f"cell://{address}", "relation": "FEEDS"})
                        if len(edges) > MAX_DEPENDENCY_EDGES:
                            raise ValueError("workbook_semantic_edge_limit_exceeded")
                    for range_obj in ranges:
                        if not range_obj.get("expanded"):
                            edges.append({"from": range_obj["objectId"], "to": f"cell://{address}", "relation": "FEEDS_RANGE"})
                else:
                    objects.append({
                        "objectId": f"cell://{address}",
                        "sheet": ws.title,
                        "cell": cell.coordinate,
                        "kind": "VALUE",
                    })
    finally:
        wb.close()

    objects.extend(range_objects[key] for key in sorted(range_objects))
    components = _strongly_connected_components(edges)
    self_edges = {edge["from"] for edge in edges if edge["from"] == edge["to"]}
    cycles = [component for component in components if len(component) > 1 or any(node in self_edges for node in component)]

    canonical_graph = {
        "workbook": path.name,
        "sheetCount": sheet_count,
        "populatedCellCount": populated,
        "formulaCellCount": formulas,
        "objectCount": len(objects),
        "dependencyEdgeCount": len(edges),
        "stronglyConnectedComponentCount": len(components),
        "cycleCount": len(cycles),
        "cycles": cycles,
        "objects": objects,
        "edges": edges,
        "limits": {
            "maxPopulatedCells": MAX_POPULATED_CELLS,
            "maxDependencyEdges": MAX_DEPENDENCY_EDGES,
            "maxRangeExpansion": MAX_RANGE_EXPANSION,
        },
        "claimBoundary": "Dependency and cycle analysis is bounded static formula-graph analysis. It does not execute Excel calculation, macros, volatile functions, external links, or prove formula correctness.",
    }
    canonical_graph["graphHash"] = sha256_bytes(canonical_json_bytes(canonical_graph))
    return canonical_graph


def write_semantic_index(workbook_path: Path) -> dict[str, Any]:
    graph = analyze_workbook(workbook_path)
    sidecar = workbook_path.with_suffix(workbook_path.suffix + ".semantics.json")
    atomic_write_text(sidecar, json.dumps(graph, indent=2, sort_keys=True) + "\n")
    return {
        "sidecar": sidecar.as_posix(),
        "graphHash": graph["graphHash"],
        "sheetCount": graph["sheetCount"],
        "populatedCellCount": graph["populatedCellCount"],
        "formulaCellCount": graph["formulaCellCount"],
        "objectCount": graph["objectCount"],
        "dependencyEdgeCount": graph["dependencyEdgeCount"],
        "cycleCount": graph["cycleCount"],
        "cycles": graph["cycles"],
    }
