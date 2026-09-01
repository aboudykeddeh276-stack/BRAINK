#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import re
import statistics
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[2]
WORKBOOK_ROOT = BASE / "workbooks"
UPLOAD_ROOT = BASE / "runtime" / "workbooks"

DATASETS = {
    "core-lattice": "CORE_LATTICE",
    "flow-territories": "FLOW_TERRITORIES",
    "genome-store": "GENOME_STORE",
    "kex-dna": "KEX_DNA",
    "substrate-grid": "SUBSTRATE_GRID",
    "system-state": "SYSTEM_STATE",
    "work-log": "WORK_LOG",
    "global-index": "GLOBAL_INDEX",
    "root-matrix": "ROOT_MATRIX",
    "benchmarks": "BENCHMARKS",
    "hyper-cores": "HYPER_CORES",
    "servers": "SERVERS",
    "storage-devices": "STORAGE_DEVICES",
    "logs": "LOGS",
}

SHEET_ALIASES = {
    "WORK_LOG": {"WORK_LOG", "WORK LOG", "WORK-LOG", "WORKLOG"},
    "GLOBAL_INDEX": {"GLOBAL_INDEX", "GLOBAL INDEX", "IL-LLM GLOBAL INDEX", "IL_LLM_GLOBAL_INDEX"},
    "ROOT_MATRIX": {"ROOT_MATRIX", "ROOT MATRIX", "ROOT-MATRIX"},
    "HYPER_CORES": {"HYPER_CORES", "HYPER CORES", "HYPER-CORES"},
    "STORAGE_DEVICES": {"STORAGE_DEVICES", "STORAGE DEVICES", "STORAGE-DEVICES"},
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _workbook_paths() -> list[Path]:
    found: list[Path] = []
    for root in (WORKBOOK_ROOT, UPLOAD_ROOT):
        if not root.exists():
            continue
        for path in sorted(root.rglob("*")):
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}:
                found.append(path)
    return found


def discover_workbooks() -> list[dict[str, Any]]:
    return [
        {
            "path": path.relative_to(BASE).as_posix(),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        for path in _workbook_paths()
    ]


def _normalise_sheet(name: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", name.upper()).strip("_")


def _normalise_header(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"column_{index + 1}"
    text = re.sub(r"[^A-Za-z0-9]+", " ", str(value)).strip()
    parts = text.split()
    if not parts:
        return f"column_{index + 1}"
    return parts[0].lower() + "".join(part[:1].upper() + part[1:] for part in parts[1:])


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _candidate_sheet_names(canonical: str) -> set[str]:
    names = {canonical, canonical.replace("_", " "), canonical.replace("_", "-")}
    names.update(SHEET_ALIASES.get(canonical, set()))
    return {_normalise_sheet(name) for name in names}


def _sheet_rows(path: Path, canonical_sheet: str) -> tuple[str, list[dict[str, Any]]] | None:
    wb = load_workbook(path, read_only=True, data_only=True)
    wanted = _candidate_sheet_names(canonical_sheet)
    selected = next((name for name in wb.sheetnames if _normalise_sheet(name) in wanted), None)
    if selected is None:
        wb.close()
        return None

    ws = wb[selected]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        wb.close()
        return selected, []

    header_row_index = 0
    for idx, row in enumerate(values[:25]):
        if sum(v is not None and str(v).strip() != "" for v in row) >= 2:
            header_row_index = idx
            break

    headers = [_normalise_header(value, idx) for idx, value in enumerate(values[header_row_index])]
    rows: list[dict[str, Any]] = []
    for row in values[header_row_index + 1:]:
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue
        record = {headers[idx]: _json_value(value) for idx, value in enumerate(row) if idx < len(headers)}
        rows.append(record)
    wb.close()
    return selected, rows


def source_not_resident(dataset: str) -> dict[str, Any]:
    return {
        "state": "SOURCE_NOT_RESIDENT",
        "dataset": dataset,
        "rows": [],
        "workbooks_discovered": discover_workbooks(),
        "claim_boundary": "The API route is resident, but no workbook containing the requested sheet was resolved in the repository/runtime workbook mounts. No rows are invented from narrative descriptions.",
    }


def dataset_response(dataset: str, *, timestamp_from: str | None = None, timestamp_to: str | None = None, operation: str | None = None) -> dict[str, Any]:
    canonical = DATASETS[dataset]
    for path in _workbook_paths():
        resolved = _sheet_rows(path, canonical)
        if resolved is None:
            continue
        sheet, rows = resolved

        if dataset in {"global-index", "logs"} and (timestamp_from or timestamp_to):
            def keep(row: dict[str, Any]) -> bool:
                value = row.get("timestamp")
                if value is None:
                    return False
                text = str(value)
                return (not timestamp_from or text >= timestamp_from) and (not timestamp_to or text <= timestamp_to)
            rows = [row for row in rows if keep(row)]

        if dataset == "benchmarks" and operation:
            rows = [row for row in rows if str(row.get("operation", "")) == operation]

        return {
            "state": "RESIDENT",
            "dataset": dataset,
            "sheet": sheet,
            "source": {
                "path": path.relative_to(BASE).as_posix(),
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            },
            "rows": rows,
            "row_count": len(rows),
            "workbooks_discovered": discover_workbooks(),
            "claim_boundary": "Rows were parsed from the named resident workbook sheet. This proves source resolution and readback, not execution of unrelated workbook triggers or external services.",
        }
    return source_not_resident(dataset)


def find_by_id(dataset: str, object_id: str) -> dict[str, Any] | None:
    response = dataset_response(dataset)
    if response["state"] != "RESIDENT":
        return None
    candidates = {"id", "coreId", "serverId", "deviceId", "identifier", "name"}
    for row in response["rows"]:
        for key in candidates:
            if key in row and str(row[key]) == object_id:
                return row
    return None


def root_matrix_statistics(rows: list[dict[str, Any]]) -> dict[str, Any]:
    values: list[float] = []
    for row in rows:
        candidate = row.get("values")
        if isinstance(candidate, (list, tuple)):
            seq = candidate
        else:
            seq = [value for key, value in row.items() if key.lower() not in {"id", "identifier"}]
        for value in seq:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                values.append(float(value))
    if not values:
        return {"count": 0, "mean": None, "min": None, "max": None, "state": "SOURCE_NOT_RESIDENT"}
    return {
        "count": len(values),
        "mean": statistics.fmean(values),
        "min": min(values),
        "max": max(values),
        "state": "COMPUTED_FROM_RESIDENT_ROWS",
    }


def _capacity_tb(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(TB|GB|PB)", value.upper())
    if not match:
        return None
    number = float(match.group(1))
    unit = match.group(2)
    return number if unit == "TB" else number / 1000 if unit == "GB" else number * 1000


def storage_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    categories: dict[str, dict[str, Any]] = {}
    for row in rows:
        category = str(row.get("type") or row.get("category") or "UNCLASSIFIED")
        slot = categories.setdefault(category, {"category": category, "count": 0, "totalCapacityTb": 0.0})
        slot["count"] += 1
        capacity = _capacity_tb(row.get("capacity"))
        if capacity is not None:
            slot["totalCapacityTb"] += capacity
    return {
        "state": "SOURCE_NOT_RESIDENT" if not rows else "COMPUTED_FROM_RESIDENT_ROWS",
        "categories": list(categories.values()),
        "source_row_count": len(rows),
    }


def system_summary(hyper_cores: list[dict[str, Any]], servers: list[dict[str, Any]], storage: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "state": "SOURCE_NOT_RESIDENT" if not (hyper_cores or servers or storage) else "COMPUTED_FROM_RESIDENT_ROWS",
        "categories": [
            {"category": "hyper_cores", "count": len(hyper_cores)},
            {"category": "servers", "count": len(servers)},
            {"category": "storage_devices", "count": len(storage)},
        ],
    }


def virtualization_metrics() -> dict[str, Any]:
    return {
        "state": "UNMEASURED",
        "cpuOverheadPct": None,
        "memoryOverheadPct": None,
        "storageOverheadPct": None,
        "networkOverheadPct": None,
        "measured_at": None,
        "claim_boundary": "No virtualization benchmark receipt is resident; numeric overhead values are not fabricated.",
    }


def activation_receipt(filename: str, payload: bytes) -> dict[str, Any]:
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    safe = Path(filename).name
    target = UPLOAD_ROOT / safe
    target.write_bytes(payload)

    sheets: list[str] = []
    parse_state = "STORED_UNPARSED"
    try:
        wb = load_workbook(target, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        parse_state = "STORED_AND_INDEXED"
    except Exception:
        parse_state = "STORED_PARSE_FAILED"

    return {
        "signal": "WORKBOOK_STORED_FOR_RESOLUTION",
        "asyncJobId": None,
        "path": target.relative_to(BASE).as_posix(),
        "bytes": len(payload),
        "sha256": hashlib.sha256(payload).hexdigest(),
        "timestamp": time.time(),
        "parse_state": parse_state,
        "sheets": sheets,
        "claim_boundary": "The receipt proves workbook bytes were accepted and, when parse_state is STORED_AND_INDEXED, that workbook metadata was successfully read. It does not prove all sheet mechanics executed.",
    }
