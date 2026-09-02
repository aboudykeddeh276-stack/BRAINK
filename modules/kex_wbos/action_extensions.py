#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import os
import tempfile
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from action_runtime import BASE, _receipt
from hardening import atomic_write_text, contained_path


def _sha(path: Path) -> str | None:
    if not path.exists():
        return None
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _find_workbook(workbook_id: str) -> Path | None:
    for root in (BASE / "workbooks", BASE / "runtime" / "workbooks"):
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and (path.stem == workbook_id or path.name == workbook_id):
                return contained_path(BASE, path)
    return None


def append_workbook_rows(workbook_id: str, table_id: str, request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"WAP-{uuid.uuid4().hex[:12]}"
    path = _find_workbook(workbook_id)
    rows = request.get("rows")
    if path is None:
        return _receipt(action_id, "BLOCKED", False, f"{workbook_id}/{table_id}", details={"error": "workbook_not_found"})
    if not isinstance(rows, list) or not rows:
        return _receipt(action_id, "FAIL", False, f"{workbook_id}/{table_id}", details={"error": "rows_required"})
    if any(not isinstance(row, dict) for row in rows):
        return _receipt(action_id, "FAIL", False, f"{workbook_id}/{table_id}", details={"error": "each_row_must_be_object"})

    wb = load_workbook(path, data_only=False)
    if table_id not in wb.sheetnames:
        return _receipt(action_id, "BLOCKED", False, f"{workbook_id}/{table_id}", details={"error": "table_not_found"})
    ws = wb[table_id]
    headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if not headers or all(v is None for v in headers):
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        headers = keys
        ws.append(headers)

    before = _sha(path)
    for row in rows:
        ws.append([row.get(str(h)) for h in headers])

    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=path.suffix, dir=str(path.parent))
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        with tmp.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(tmp, path)
        try:
            dfd = os.open(path.parent, os.O_RDONLY)
            try:
                os.fsync(dfd)
            finally:
                os.close(dfd)
        except OSError:
            pass
    finally:
        tmp.unlink(missing_ok=True)

    after = _sha(path)
    if before == after:
        return _receipt(action_id, "FAIL", False, f"{workbook_id}/{table_id}", before=before, after=after, details={"error": "mutation_hash_unchanged"})
    return _receipt(action_id, "MUTATED", True, f"{workbook_id}/{table_id}", before=before, after=after, details={"path": path.relative_to(BASE).as_posix(), "rowsAppended": len(rows)})


def commit_braink_migration(request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"MIG-{uuid.uuid4().hex[:12]}"
    source = Path(str(request.get("sourcePath", ""))).expanduser().resolve()
    target_name = Path(str(request.get("targetWorkbook", "BRAINK_KERNEL_WORKBOOK.xlsx"))).name
    target = contained_path(BASE, BASE / "runtime" / "workbooks" / target_name)
    if not source.exists():
        return _receipt(action_id, "BLOCKED", False, target_name, details={"error": "source_path_not_found", "sourcePath": str(source)})

    allow_external = os.getenv("KEX_ALLOW_EXTERNAL_MIGRATION_SOURCE", "false").lower() == "true"
    if not allow_external:
        try:
            contained_path(BASE, source)
        except ValueError:
            return _receipt(action_id, "BLOCKED", False, target_name, details={
                "error": "external_migration_source_not_authorized",
                "sourcePath": str(source),
                "requiredSetting": "KEX_ALLOW_EXTERNAL_MIGRATION_SOURCE=true",
            })

    files = [p for p in source.rglob("*") if p.is_file()] if source.is_dir() else [source]
    manifest = {str(p): hashlib.sha256(p.read_bytes()).hexdigest() for p in files}
    if request.get("requireHashManifest", True) and not manifest:
        return _receipt(action_id, "BLOCKED", False, target_name, details={"error": "hash_manifest_empty"})
    target.parent.mkdir(parents=True, exist_ok=True)
    manifest_path = target.with_suffix(".manifest.json")
    before = _sha(manifest_path)
    manifest_payload = {
        "sourcePath": str(source),
        "targetWorkbook": target_name,
        "files": manifest,
        "migrateTranscripts": bool(request.get("migrateTranscripts", True)),
        "migrateArtifacts": bool(request.get("migrateArtifacts", True)),
    }
    atomic_write_text(manifest_path, json.dumps(manifest_payload, indent=2, sort_keys=True))
    after = _sha(manifest_path)
    return _receipt(action_id, "MUTATED", True, target_name, before=before, after=after, details={"manifest": manifest_path.relative_to(BASE).as_posix(), "fileCount": len(files), "claimBoundary": "This commits a migration manifest and source hash set. It does not claim semantic ingestion into workbook cells unless a workbook writer performs that separate mutation."})
