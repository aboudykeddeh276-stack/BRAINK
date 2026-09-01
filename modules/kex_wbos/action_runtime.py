#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import os
import subprocess
import time
import urllib.error
import urllib.request
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from capabilities import verify_capability
from casepath_management import managed_dispatch
from hardening import append_jsonl_fsync, atomic_write_text, contained_path
from idempotency import IdempotencyRegistry
from network_policy import readback_url_allowed, runtime_auth_allowed
from object_store import ContentAddressedStore

BASE = Path(__file__).resolve().parents[2]
RUNTIME = BASE / "runtime"
LEGACY_ACTION_LEDGER = BASE / "reports" / "kex-wbos" / "action-ledger.jsonl"
ACTION_LEDGER = BASE / "reports" / "kex-wbos" / "action-ledger-v2.jsonl"
SOURCE_ROOT = RUNTIME / "sources"
DISPATCH_ROOT = RUNTIME / "casepath-dispatch"
PROOF_ROOT = BASE / "reports" / "kex-wbos"
OBJECT_STORE = ContentAddressedStore(RUNTIME / "objects")
IDEMPOTENCY = IdempotencyRegistry(RUNTIME / "idempotency" / "action-registry-v1.json")

EXTERNAL_ACTIONS = {
    "PUBLIC_DNS": "DNS_PROVIDER_ADAPTER",
    "PUBLIC_TLS": "ACME_OR_TLS_ADAPTER",
    "ROUTER_FIREWALL": "ROUTER_ADMIN_ADAPTER",
    "BITCOIN_LIVE_SUBMISSION": "BITCOIN_RPC_ADAPTER",
    "DRIVE_WRITEBACK": "GOOGLE_DRIVE_ADAPTER",
    "LIVE_PUBLIC_DEPLOYMENT": "DEPLOYMENT_ADAPTER",
}


class _NoRedirect(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # noqa: ANN001
        return None


NO_REDIRECT_OPENER = urllib.request.build_opener(_NoRedirect())


def _sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _now() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


def _receipt(action_id: str, status: str, mutated: bool, target: str, *, before: str | None = None, after: str | None = None, external_readback: bool = False, details: dict[str, Any] | None = None) -> dict[str, Any]:
    receipt = {
        "ledgerVersion": 2,
        "status": status,
        "actionId": action_id,
        "mutated": mutated,
        "target": target,
        "receiptId": f"KEXR-{uuid.uuid4().hex[:16]}",
        "beforeHash": before,
        "afterHash": after,
        "proofLedgerRow": None,
        "parentReceiptHash": None,
        "receiptHash": None,
        "externalReadback": external_readback,
        "timestamp": _now(),
        "details": details or {},
    }
    _, persisted = append_jsonl_fsync(
        ACTION_LEDGER,
        receipt,
        row_field="proofLedgerRow",
        hash_field="receiptHash",
        parent_hash_field="parentReceiptHash",
    )
    return persisted


def _capability_check(request: dict[str, Any], action_type: str, target: str) -> tuple[bool, dict[str, Any] | None]:
    if os.getenv("KEX_REQUIRE_SCOPED_CAPABILITIES", "false").lower() != "true":
        return True, None
    secret = os.getenv("KEX_CAPABILITY_SECRET", "")
    if not secret:
        return False, {"error": "capability_secret_not_configured"}
    return verify_capability(secret, str(request.get("capability", "")), action=action_type, target=target)


def _execute_action_once(request: dict[str, Any]) -> dict[str, Any]:
    action_type = str(request.get("actionType", "")).upper()
    target = str(request.get("target", ""))
    action_id = f"ACT-{uuid.uuid4().hex[:12]}"
    if not request.get("authority"):
        return _receipt(action_id, "FAIL", False, target, details={"error": "authority_required"})
    if not action_type or not target:
        return _receipt(action_id, "FAIL", False, target, details={"error": "actionType_and_target_required"})

    capability_ok, capability_detail = _capability_check(request, action_type, target)
    if not capability_ok:
        return _receipt(action_id, "BLOCKED", False, target, details={"error": "capability_denied", "capability": capability_detail})

    if action_type in EXTERNAL_ACTIONS:
        adapter = request.get("controlRoute") or os.getenv(f"KEX_{EXTERNAL_ACTIONS[action_type]}")
        if not adapter:
            return _receipt(action_id, "BLOCKED", False, target, details={"requiredAdapter": EXTERNAL_ACTIONS[action_type], "claimBoundary": "External mutation is not claimed without a bound actuator and receipt."})
    if action_type == "SOURCE_INGEST":
        payload = request.get("payload", {})
        return ingest_source({"authority": request["authority"], "sourceText": payload.get("sourceText", ""), "sourceFormat": payload.get("sourceFormat", "markdown"), "target": target})
    if action_type == "CASEPATH_DISPATCH":
        payload = request.get("payload", {})
        return dispatch_casepath({
            "authority": request["authority"],
            "packetId": payload.get("packetId", action_id),
            "activeTarget": target,
            "processId": payload.get("processId", "CASEPATH-PROC-001"),
            "actionQueue": payload.get("actionQueue", []),
            "proofTarget": payload.get("proofTarget"),
        })
    if action_type == "PROOF_LEDGER_WRITE":
        return write_proof({"authority": request["authority"], "eventType": request.get("payload", {}).get("eventType", action_type), "payload": request.get("payload", {}), "targetLedger": target})
    return _receipt(action_id, "ARMED", False, target, details={"actionType": action_type, "controlRoute": request.get("controlRoute"), "capability": capability_detail, "claimBoundary": "Action class is resident but no local executor was selected for this request."})


def execute_action(request: dict[str, Any]) -> dict[str, Any]:
    key = str(request.get("idempotencyKey", "")).strip()
    if not key:
        return _execute_action_once(request)

    begin = IDEMPOTENCY.begin(key, request)
    state = begin.get("state")
    target = str(request.get("target", ""))
    if state == "REPLAY":
        receipt = begin.get("receipt")
        if isinstance(receipt, dict):
            return receipt
        return _receipt(f"IDEM-{uuid.uuid4().hex[:12]}", "FAIL", False, target, details={"error": "idempotency_replay_receipt_missing", "idempotencyKey": key})
    if state == "CONFLICT":
        return _receipt(f"IDEM-{uuid.uuid4().hex[:12]}", "BLOCKED", False, target, details={"error": "idempotency_key_conflict", "idempotencyKey": key, **begin})
    if state == "INFLIGHT":
        return _receipt(f"IDEM-{uuid.uuid4().hex[:12]}", "BLOCKED", False, target, details={"error": "idempotency_outcome_ambiguous", "idempotencyKey": key, "claimBoundary": "A prior execution reserved this key but did not commit a terminal receipt. Automatic replay is suppressed to avoid duplicate mutation; reconcile the original effect before retrying."})
    if state != "NEW":
        return _receipt(f"IDEM-{uuid.uuid4().hex[:12]}", "FAIL", False, target, details={"error": "invalid_idempotency_key", "idempotencyKey": key})

    receipt = _execute_action_once(request)
    IDEMPOTENCY.complete(key, str(begin["requestHash"]), receipt)
    return receipt


def ingest_source(request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"SRC-{uuid.uuid4().hex[:12]}"
    text = request.get("sourceText")
    target = str(request.get("target", "KEX_RUNTIME_MODEL"))
    if not isinstance(text, str) or not text:
        return _receipt(action_id, "FAIL", False, target, details={"error": "sourceText_required"})
    SOURCE_ROOT.mkdir(parents=True, exist_ok=True)
    source_format = request.get("sourceFormat", "text")
    ext = {"markdown": "md", "json": "json", "html": "html", "text": "txt"}.get(source_format, "txt")
    raw = text.encode("utf-8")
    object_receipt = OBJECT_STORE.put_bytes(
        raw,
        media_type={"markdown": "text/markdown", "json": "application/json", "html": "text/html", "text": "text/plain"}.get(source_format, "text/plain"),
        metadata={"target": target, "authority": request.get("authority"), "ingestActionId": action_id},
    )
    OBJECT_STORE.bind_ref(f"source-{action_id}", object_receipt["objectId"])
    path = contained_path(SOURCE_ROOT, SOURCE_ROOT / f"{action_id}.{ext}")
    before = _sha(path.read_bytes()) if path.exists() else None
    atomic_write_text(path, text)
    after = _sha(path.read_bytes())
    if after != object_receipt["sha256"]:
        return _receipt(action_id, "FAIL", False, target, before=before, after=after, details={"error": "source_carrier_hash_mismatch", "objectId": object_receipt["objectId"]})
    return _receipt(
        action_id,
        "MUTATED",
        True,
        target,
        before=before,
        after=after,
        details={
            "path": path.relative_to(BASE).as_posix(),
            "bytes": path.stat().st_size,
            "objectId": object_receipt["objectId"],
            "contentAddress": object_receipt["sha256"],
            "identityBoundary": "Filesystem path is a carrier; sha256 object ID is the immutable content identity.",
        },
    )


def dispatch_casepath(request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"DSP-{uuid.uuid4().hex[:12]}"
    return managed_dispatch(base=BASE, dispatch_root=DISPATCH_ROOT, request=request, receipt=_receipt, now=_now, sha=_sha, action_id=action_id)


def write_proof(request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"PRF-{uuid.uuid4().hex[:12]}"
    event = {"ts": time.time(), "authority": request.get("authority"), "eventType": request.get("eventType"), "payload": request.get("payload", {}), "targetLedger": request.get("targetLedger", "KEX_ACTION_LEDGER")}
    path = PROOF_ROOT / "action-proof-ledger.jsonl"
    before = _sha(path.read_bytes()) if path.exists() else None
    row, persisted_event = append_jsonl_fsync(path, event)
    after = _sha(path.read_bytes())
    return _receipt(action_id, "MUTATED", True, str(event["targetLedger"]), before=before, after=after, details={"event": persisted_event, "proofEventRow": row})


def readback_runtime(request: dict[str, Any]) -> dict[str, Any]:
    target = str(request.get("target", ""))
    expected_text = request.get("expectedText")
    if target.startswith("http://") or target.startswith("https://"):
        allowed, policy = readback_url_allowed(target)
        if not allowed:
            return {"status": "FAIL", "target": target, "matched": False, "observedState": "READBACK_DESTINATION_BLOCKED", "proofHash": None, "policy": policy}
        try:
            req = urllib.request.Request(target)
            token = os.getenv("KEX_BEARER_TOKEN")
            if token and request.get("sendRuntimeAuth", True) and runtime_auth_allowed(target):
                req.add_header("Authorization", f"Bearer {token}")
            try:
                response = NO_REDIRECT_OPENER.open(req, timeout=10)
            except urllib.error.HTTPError as exc:
                if 300 <= exc.code < 400:
                    return {"status": "FAIL", "target": target, "matched": False, "observedState": "REDIRECT_BLOCKED", "proofHash": None, "policy": policy, "location": exc.headers.get("Location")}
                raise
            with response:
                body = response.read()
                text = body.decode("utf-8", errors="replace")
                matched = response.status < 400 and (expected_text is None or expected_text in text)
                return {"status": "VERIFIED" if matched else "FAIL", "target": target, "matched": matched, "observedState": f"HTTP_{response.status}", "proofHash": _sha(body), "policy": policy, "runtimeAuthSent": bool(token and request.get("sendRuntimeAuth", True) and runtime_auth_allowed(target))}
        except Exception as exc:
            return {"status": "FAIL", "target": target, "matched": False, "observedState": type(exc).__name__, "proofHash": None, "policy": policy}
    candidate = Path(target)
    if not candidate.is_absolute():
        candidate = BASE / candidate
    try:
        path = contained_path(BASE, candidate)
    except ValueError:
        return {"status": "FAIL", "target": target, "matched": False, "observedState": "PATH_OUTSIDE_RUNTIME_ROOT", "proofHash": None}
    if path.exists() and path.is_file():
        raw = path.read_bytes()
        text = raw.decode("utf-8", errors="replace")
        matched = expected_text is None or expected_text in text
        return {"status": "VERIFIED" if matched else "FAIL", "target": target, "matched": matched, "observedState": "FILE_PRESENT", "proofHash": _sha(raw)}
    return {"status": "FAIL", "target": target, "matched": False, "observedState": "NOT_FOUND", "proofHash": None}


def launch_runtime(request: dict[str, Any]) -> dict[str, Any]:
    action_id = f"RUN-{uuid.uuid4().hex[:12]}"
    runtime_id = str(request.get("runtimeId", ""))
    route = str(request.get("commandRoute", ""))
    if not runtime_id or not route:
        return _receipt(action_id, "FAIL", False, runtime_id, details={"error": "runtimeId_and_commandRoute_required"})
    allowed = {"kex-wbos-self-test": [os.sys.executable, str(BASE / "scripts" / "kex-ci" / "test_wbos_api.py")]}
    command = allowed.get(route)
    if not command:
        return _receipt(action_id, "BLOCKED", False, runtime_id, details={"error": "commandRoute_not_bound", "allowedRoutes": sorted(allowed)})
    proc = subprocess.run(command, cwd=BASE, capture_output=True, text=True, timeout=120, shell=False)
    output = (proc.stdout + "\n" + proc.stderr).encode()
    return _receipt(action_id, "VERIFIED" if proc.returncode == 0 else "FAIL", False, runtime_id, after=_sha(output), details={"returnCode": proc.returncode, "stdout": proc.stdout[-4000:], "stderr": proc.stderr[-4000:]})


def read_workbook_table(workbook_id: str, table_id: str) -> tuple[int, dict[str, Any]]:
    roots = [BASE / "workbooks", BASE / "runtime" / "workbooks"]
    candidates: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        for pattern in ("*.xlsx", "*.xlsm"):
            for candidate in root.rglob(pattern):
                try:
                    safe = contained_path(BASE, candidate)
                except ValueError:
                    continue
                if safe.stem == workbook_id or safe.name == workbook_id:
                    candidates.append(safe)
    if not candidates:
        return 404, {"error": "workbook_not_found", "workbookId": workbook_id, "tableId": table_id}
    workbook = candidates[0]
    wb = load_workbook(workbook, data_only=False, read_only=True)
    try:
        if table_id not in wb.sheetnames:
            return 404, {"error": "table_not_found", "workbookId": workbook_id, "tableId": table_id}
        ws = wb[table_id]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            rows: list[dict[str, Any]] = []
        else:
            headers = [str(v) if v is not None else f"column_{i+1}" for i, v in enumerate(values[0])]
            rows = [{headers[i]: row[i] if i < len(row) else None for i in range(len(headers))} for row in values[1:] if any(v is not None for v in row)]
        return 200, {"workbookId": workbook_id, "tableId": table_id, "rowCount": len(rows), "rows": rows}
    finally:
        wb.close()
